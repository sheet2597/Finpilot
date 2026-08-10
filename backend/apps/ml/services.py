"""Orchestration layer for the ML module.

Provides:
- resolve_company_scope: helper to resolve company access
- train_single_model: trains a single model type on demand
- Report builders: XLSX/PDF downloads for ML predictions
  - build_duplicate_transaction_report (fraud/anomaly detection)
  - build_expense_forecast_report
  - build_tax_prediction_report
  - build_compliance_risk_report
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

from apps.companies.services import get_accessible_company
from apps.documents.services import get_user_company_ids
from apps.mongo import connection as mongo_connection

from . import predict as pred
from . import train
from . import utils as ml_utils
from .features import InsufficientDataError
from .predict import ModelNotTrainedError

HEADER_FILL = "1E293B"


def resolve_company_scope(user, company_id=None):
    if company_id:
        db = mongo_connection.get_db()
        company, _is_owner = get_accessible_company(db, user, company_id)
        return [company["_id"]], company
    return get_user_company_ids(user), None


# ---------------------------------------------------------------------------
# Financial Health Score
# ---------------------------------------------------------------------------

def compute_financial_health_score(user, company_id=None, summary=None):
    """Computes a holistic financial health score (0-100) based on transaction
    data, compliance signals, and ML predictions. Returns score, label,
    strengths, and weaknesses for dashboard display."""
    from apps.transactions.services import get_dashboard_summary

    def _r(v):
        try:
            return float(v or 0)
        except (TypeError, ValueError):
            return 0.0

    try:
        if summary is None:
            summary = get_dashboard_summary(user, company_id)
    except Exception:
        summary = {}

    fi = summary.get("financial_intelligence", {})
    income = _r(summary.get("total_income", 0))
    net_profit = _r(summary.get("net_profit", 0))

    # Component 1: Profitability (0-30)
    profit_margin = (net_profit / income * 100) if income > 0 else 0
    profitability_score = min(30, max(0, profit_margin * 0.6))

    # Component 2: Cash-flow stability (0-25)
    trend = fi.get("cash_flow_trend", [])
    if len(trend) >= 3:
        recent_cf = [t.get("cash_flow", 0) for t in trend[-3:]]
        positive_months = sum(1 for v in recent_cf if v >= 0)
        cash_flow_score = (positive_months / 3) * 25
    else:
        cash_flow_score = 12.5  # neutral if no data

    # Component 3: Expense control (0-25)
    expense_growth = _r(fi.get("expense_growth", 0))
    income_growth = _r(fi.get("income_growth", 0))
    if income_growth > 0 and expense_growth <= income_growth:
        expense_score = 25
    elif expense_growth <= 5:
        expense_score = 20
    elif expense_growth <= 15:
        expense_score = 12
    else:
        expense_score = 5

    # Component 4: Revenue consistency (0-20)
    monthly = summary.get("monthly_trend", [])
    if len(monthly) >= 3:
        income_vals = [m.get("income", 0) for m in monthly[-6:]]
        avg = sum(income_vals) / len(income_vals) if income_vals else 0
        variance = sum((v - avg) ** 2 for v in income_vals) / len(income_vals) if income_vals else 0
        cv = (variance ** 0.5 / avg) if avg > 0 else 1
        revenue_score = max(0, 20 - cv * 10)
    else:
        revenue_score = 10

    overall_score = round(ml_utils.round2(profitability_score + cash_flow_score + expense_score + revenue_score))
    overall_score = max(0, min(100, overall_score))

    if overall_score >= 75:
        label, color = "Excellent", "green"
    elif overall_score >= 55:
        label, color = "Good", "blue"
    elif overall_score >= 35:
        label, color = "Fair", "yellow"
    else:
        label, color = "Needs Attention", "red"

    strengths = []
    weaknesses = []
    if profit_margin > 10:
        strengths.append("Strong profit margins")
    elif profit_margin < 0:
        weaknesses.append("Negative profitability")
    if cash_flow_score >= 20:
        strengths.append("Stable cash flow")
    elif cash_flow_score < 10:
        weaknesses.append("Volatile cash flow")
    if expense_score >= 20:
        strengths.append("Controlled expenses")
    elif expense_score < 10:
        weaknesses.append("High expense growth")
    if revenue_score >= 16:
        strengths.append("Consistent revenue")
    elif revenue_score < 8:
        weaknesses.append("Inconsistent revenue")

    return {
        "overall_score": overall_score,
        "label": label,
        "color": color,
        "strengths": strengths,
        "weaknesses": weaknesses,
        "suggestions": weaknesses + strengths,
        "components": {
            "profitability": round(profitability_score),
            "cash_flow": round(cash_flow_score),
            "expense_control": round(expense_score),
            "revenue_consistency": round(revenue_score),
        },
    }


# ---------------------------------------------------------------------------
# AI Dashboard (aggregated ML predictions for executive view)
# ---------------------------------------------------------------------------

def get_ml_dashboard(user, company_id=None):
    """Returns aggregated ML predictions for the ML dashboard.
    Combines compliance risk, fraud/duplicate detection, and expense forecast
    into a single unified payload. Raises ModelNotTrainedError or
    InsufficientDataError if models are not yet trained (caller uses _safe())."""
    company_ids, _ = resolve_company_scope(user, company_id)

    # Compliance risk
    try:
        compliance_results = pred.predict_compliance_risk(user, company_ids)
        if compliance_results:
            worst = max(compliance_results, key=lambda r: r.get("confidence", 0))
            compliance_prediction = {
                "risk_level": worst.get("risk_level", "unknown"),
                "confidence": worst.get("confidence", 0),
            }
        else:
            compliance_prediction = {"risk_level": "low", "confidence": 0.5}
    except (ModelNotTrainedError, InsufficientDataError):
        compliance_prediction = {"available": False, "reason": "Model not trained yet."}

    # Risk alerts from fraud/anomaly detection
    risk_alerts = []
    try:
        anomalies = pred.predict_fraud_anomalies(user, company_id)
        for a in anomalies[:5]:
            if a.get("risk_level") in ("high", "medium"):
                risk_alerts.append({
                    "type": "duplicate_transaction",
                    "message": f"Transaction {a['transaction_id']}: {a['risk_level']} risk (score {a['risk_score']}/100)",
                    "risk_level": a["risk_level"],
                })
    except (ModelNotTrainedError, InsufficientDataError):
        pass

    # Expense forecast insights
    insights = []
    try:
        forecast, entry = pred.predict_expense_forecast(user, periods_ahead=3)
        if forecast:
            next_month = forecast[0]
            insights.append({
                "type": "expense_forecast",
                "message": f"Projected expenses for next month: ₹{next_month['predicted_total']:.2f}",
            })
    except (ModelNotTrainedError, InsufficientDataError):
        pass

    return {
        "compliance_prediction": compliance_prediction,
        "risk_alerts": risk_alerts,
        "insights": insights,
        "model_count": len(ml_utils.list_model_registry(user.id)),
    }


# ---------------------------------------------------------------------------
# Model management
# ---------------------------------------------------------------------------

def train_single_model(user, model_type, company_id=None, is_demo=False):
    company_ids, _ = resolve_company_scope(user, company_id)
    dispatch = {
        ml_utils.MODEL_EXPENSE_CATEGORIZATION: lambda: train.train_expense_categorization(user, company_id),
        ml_utils.MODEL_FRAUD_DETECTION: lambda: train.train_fraud_detection(user, company_id),
        ml_utils.MODEL_COMPLIANCE_RISK: lambda: train.train_compliance_risk(user, company_ids, is_demo=is_demo),
        ml_utils.MODEL_TAX_LIABILITY: lambda: train.train_tax_liability_forecast(user, company_ids),
        ml_utils.MODEL_EXPENSE_FORECAST: lambda: train.train_expense_forecast(user, company_ids),
    }
    if model_type not in dispatch:
        raise ValueError(f"Unknown model_type '{model_type}'.")
    return dispatch[model_type]()


# ---------------------------------------------------------------------------
# Reports (XLSX / PDF)
# ---------------------------------------------------------------------------

def _xlsx_from_rows(title, headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]
    sheet.append(headers)
    fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for row in rows:
        sheet.append(row)
    for column_cells in sheet.columns:
        length = max((len(str(c.value)) for c in column_cells if c.value is not None), default=10)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 40)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _pdf_from_rows(title, headers, rows, subtitle=None):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"])]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(Spacer(1, 12))

    table_data = [headers] + [[str(c) for c in row] for row in rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
    ]))
    elements.append(table)
    doc.build(elements)
    return buffer.getvalue()


def _render_report(title, headers, rows, fmt, subtitle=None):
    if fmt == "pdf":
        return _pdf_from_rows(title, headers, rows, subtitle=subtitle), "application/pdf", "pdf"
    return _xlsx_from_rows(title, headers, rows), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"


def build_duplicate_transaction_report(user, company_id, fmt):
    """Report on anomalous / duplicate transactions detected by IsolationForest."""
    try:
        anomalies = pred.predict_fraud_anomalies(user, company_id)
    except ModelNotTrainedError:
        anomalies = []
    headers = ["Transaction ID", "Risk Level", "Risk Score", "Amount", "Explanation"]
    rows = [[a["transaction_id"], a["risk_level"], a["risk_score"], a["amount"], pred.explain_fraud(a)] for a in anomalies]
    return _render_report("Duplicate Transaction Detection Report", headers, rows, fmt)


def build_expense_forecast_report(user, fmt):
    try:
        forecast, _ = pred.predict_expense_forecast(user)
    except ModelNotTrainedError:
        forecast, _ = [], {"metrics": {}}
    headers = ["Year", "Month", "Predicted Total", "Lower Bound", "Upper Bound"]
    rows = [[f["year"], f["month"], f["predicted_total"], f["lower_bound"], f["upper_bound"]] for f in forecast]
    return _render_report("Expense Forecast Report", headers, rows, fmt)


def build_tax_prediction_report(user, fmt):
    try:
        forecast, _ = pred.predict_tax_liability(user)
    except ModelNotTrainedError:
        forecast, _ = [], {"metrics": {}}
    headers = ["Year", "Month", "Predicted Tax Liability", "Lower Bound", "Upper Bound"]
    rows = [[f["year"], f["month"], f["predicted_total"], f["lower_bound"], f["upper_bound"]] for f in forecast]
    return _render_report("Tax Liability Prediction Report", headers, rows, fmt)


def build_compliance_risk_report(user, company_id, fmt):
    company_ids, _ = resolve_company_scope(user, company_id)
    try:
        results = pred.predict_compliance_risk(user, company_ids)
    except ModelNotTrainedError:
        results = []
    headers = ["Company ID", "Risk Level", "Confidence", "Missing GST Ratio", "Missing TDS Ratio", "Duplicate Count"]
    rows = [
        [
            r["company_id"], r["risk_level"], f"{round(r['confidence']*100,1)}%",
            r["features"].get("missing_gst_ratio", ""),
            r["features"].get("missing_tds_ratio", ""),
            r["features"].get("duplicate_invoice_count", ""),
        ]
        for r in results
    ]
    return _render_report("Compliance Risk Report", headers, rows, fmt)
