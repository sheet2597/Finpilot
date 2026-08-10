import csv
import difflib
import io
import re
from collections import defaultdict
from datetime import date as date_cls, timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from rest_framework.exceptions import NotFound, PermissionDenied, ValidationError

from apps.companies.services import get_accessible_company
from apps.documents.services import get_user_company_ids
from apps.mongo import connection as mongo_connection
from apps.mongo.utils import get_or_404, log_activity, now, paginate, serialize, to_object_id

DEFAULT_CATEGORIES = [
    {"name": "Sales", "type": "income", "color": "#22C55E", "icon": "trending-up"},
    {"name": "Service Revenue", "type": "income", "color": "#10B981", "icon": "briefcase"},
    {"name": "Interest Income", "type": "income", "color": "#14B8A6", "icon": "percent"},
    {"name": "Other Income", "type": "income", "color": "#06B6D4", "icon": "plus-circle"},
    {"name": "Office Supplies", "type": "expense", "color": "#F97316", "icon": "package"},
    {"name": "Rent", "type": "expense", "color": "#EF4444", "icon": "home"},
    {"name": "Utilities", "type": "expense", "color": "#F59E0B", "icon": "zap"},
    {"name": "Salaries", "type": "expense", "color": "#8B5CF6", "icon": "users"},
    {"name": "Travel", "type": "expense", "color": "#EC4899", "icon": "map-pin"},
    {"name": "Marketing", "type": "expense", "color": "#3B82F6", "icon": "megaphone"},
    {"name": "Professional Fees", "type": "expense", "color": "#6366F1", "icon": "file-text"},
    {"name": "Bank Charges", "type": "expense", "color": "#64748B", "icon": "credit-card"},
    {"name": "Other Expense", "type": "expense", "color": "#94A3B8", "icon": "more-horizontal"},
]


# ---------------------------------------------------------------------------
# Categories
# ---------------------------------------------------------------------------

def ensure_default_categories(db):
    if db.transaction_categories.count_documents({"is_default": True}) > 0:
        return
    docs = [{**c, "is_default": True, "owner_id": None, "description": "", "created_at": now()} for c in DEFAULT_CATEGORIES]
    db.transaction_categories.insert_many(docs)


def list_categories(user, category_type=None):
    db = mongo_connection.get_db()
    ensure_default_categories(db)
    query = {"$or": [{"is_default": True}, {"owner_id": str(user.id)}]}
    if category_type:
        query = {"$and": [query, {"type": category_type}]}
    categories = list(db.transaction_categories.find(query).sort([("is_default", -1), ("name", 1)]))
    return serialize(categories)


def create_category(user, data):
    db = mongo_connection.get_db()
    existing = db.transaction_categories.find_one({
        "name": {"$regex": f"^{data['name']}$", "$options": "i"},
        "type": data["type"],
        "$or": [{"is_default": True}, {"owner_id": str(user.id)}],
    })
    if existing:
        raise ValidationError({"name": "A category with this name already exists."})

    doc = {**data, "is_default": False, "owner_id": str(user.id), "created_at": now()}
    result = db.transaction_categories.insert_one(doc)
    doc["_id"] = result.inserted_id
    return serialize(doc)


def delete_category(user, category_id):
    db = mongo_connection.get_db()
    oid = to_object_id(category_id, field="category_id")
    category = get_or_404(db.transaction_categories, {"_id": oid}, "Category not found.")
    if category.get("is_default"):
        raise PermissionDenied("Default categories cannot be deleted.")
    if category.get("owner_id") != str(user.id):
        raise PermissionDenied("You can only delete your own categories.")
    if db.transactions.count_documents({"category_id": oid, "status": {"$ne": "deleted"}}) > 0:
        raise ValidationError("This category is in use by existing transactions and cannot be deleted.")
    db.transaction_categories.delete_one({"_id": oid})


# ---------------------------------------------------------------------------
# Vendors / Customers (shared implementation - identical shape per the brief)
# ---------------------------------------------------------------------------

def _party_collection(db, party_type):
    return db.vendors if party_type == "vendor" else db.customers


def list_parties(user, party_type, params):
    db = mongo_connection.get_db()
    collection = _party_collection(db, party_type)
    query = {"owner_id": str(user.id), "is_deleted": False}

    search = params.get("search")
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"gst_number": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
        ]

    return paginate(
        collection, query,
        page=params.get("page", 1), page_size=params.get("page_size", 10),
        sort_field="name", sort_dir=1,
    )


def create_party(user, party_type, data):
    db = mongo_connection.get_db()
    collection = _party_collection(db, party_type)
    doc = {
        **data, "owner_id": str(user.id), "is_deleted": False,
        "created_at": now(), "updated_at": now(),
    }
    result = collection.insert_one(doc)
    doc["_id"] = result.inserted_id
    return serialize(doc)


def get_party_detail(user, party_type, party_id):
    db = mongo_connection.get_db()
    collection = _party_collection(db, party_type)
    oid = to_object_id(party_id, field=f"{party_type}_id")
    party = get_or_404(collection, {"_id": oid, "owner_id": str(user.id), "is_deleted": False}, f"{party_type.title()} not found.")

    field = f"{party_type}_id"
    history = list(
        db.transactions.find({field: oid, "status": {"$ne": "deleted"}}).sort("date", -1).limit(50)
    )
    result = serialize(party)
    result["transaction_history"] = serialize(history)
    return result


def update_party(user, party_type, party_id, data):
    db = mongo_connection.get_db()
    collection = _party_collection(db, party_type)
    oid = to_object_id(party_id, field=f"{party_type}_id")
    get_or_404(collection, {"_id": oid, "owner_id": str(user.id), "is_deleted": False}, f"{party_type.title()} not found.")
    collection.update_one({"_id": oid}, {"$set": {**data, "updated_at": now()}})
    return serialize(collection.find_one({"_id": oid}))


def delete_party(user, party_type, party_id):
    db = mongo_connection.get_db()
    collection = _party_collection(db, party_type)
    oid = to_object_id(party_id, field=f"{party_type}_id")
    get_or_404(collection, {"_id": oid, "owner_id": str(user.id), "is_deleted": False}, f"{party_type.title()} not found.")
    collection.update_one({"_id": oid}, {"$set": {"is_deleted": True, "updated_at": now()}})


# ---------------------------------------------------------------------------
# Transactions - access + lookups
# ---------------------------------------------------------------------------

def _validate_category(db, user, category_id):
    """Category must be a default category or one owned by this user —
    never another user's private category (security fix)."""
    oid = to_object_id(category_id, field="category_id")
    get_or_404(
        db.transaction_categories,
        {"_id": oid, "$or": [{"is_default": True}, {"owner_id": str(user.id)}]},
        "Category not found.",
    )
    return oid


def _validate_vendor(db, user, vendor_id):
    oid = to_object_id(vendor_id, field="vendor_id")
    get_or_404(db.vendors, {"_id": oid, "owner_id": str(user.id), "is_deleted": False}, "Vendor not found.")
    return oid


def _validate_customer(db, user, customer_id):
    oid = to_object_id(customer_id, field="customer_id")
    get_or_404(db.customers, {"_id": oid, "owner_id": str(user.id), "is_deleted": False}, "Customer not found.")
    return oid


def _validate_document(db, company, document_id):
    oid = to_object_id(document_id, field="document_id")
    get_or_404(db.documents, {"_id": oid, "company_id": company["_id"]}, "Linked document not found for this company.")
    return oid


def _resolve_references(db, user, data):
    """Validates company/category/vendor/customer/document ids and returns
    the resolved ObjectIds, raising 404/403 for anything inaccessible."""
    company, _ = get_accessible_company(db, user, data["company_id"])

    category_oid = _validate_category(db, user, data["category_id"])
    # Category<->transaction-type is intentionally advisory, not enforced:
    # e.g. a "refund" or "adjustment" transaction may reasonably use either
    # an income or expense category. We only require the category to be
    # a default one or one owned by this user (see _validate_category).

    vendor_oid = _validate_vendor(db, user, data["vendor_id"]) if data.get("vendor_id") else None
    customer_oid = _validate_customer(db, user, data["customer_id"]) if data.get("customer_id") else None
    document_oid = _validate_document(db, company, data["document_id"]) if data.get("document_id") else None

    return company, category_oid, vendor_oid, customer_oid, document_oid


def get_accessible_transaction(db, user, transaction_id):
    oid = to_object_id(transaction_id, field="transaction_id")
    transaction = get_or_404(db.transactions, {"_id": oid, "status": {"$ne": "deleted"}}, "Transaction not found.")
    company, is_owner = get_accessible_company(db, user, str(transaction["company_id"]))
    return transaction, company, is_owner


# ---------------------------------------------------------------------------
# Transactions - CRUD
# ---------------------------------------------------------------------------

DUPLICATE_DATE_WINDOW_DAYS = 3
DESCRIPTION_SIMILARITY_THRESHOLD = 0.85


def find_possible_duplicates(db, company_id, doc, exclude_id=None):
    """Multi-signal duplicate check: amount+date, reference number, invoice
    number, linked document, vendor+amount, and description similarity.
    Returns a list of {id, reasons, amount, date} — purely informational,
    nothing here ever deletes or blocks a save.
    """
    strong_id_clauses = []
    if doc.get("reference_number"):
        strong_id_clauses.append({"reference_number": doc["reference_number"]})
    if doc.get("invoice_number"):
        strong_id_clauses.append({"invoice_number": doc["invoice_number"]})
    if doc.get("document_id"):
        strong_id_clauses.append({"document_id": doc["document_id"]})

    try:
        center = date_cls.fromisoformat(doc["date"])
        date_range = [
            (center + timedelta(days=d)).isoformat()
            for d in range(-DUPLICATE_DATE_WINDOW_DAYS, DUPLICATE_DATE_WINDOW_DAYS + 1)
        ]
    except (ValueError, TypeError):
        date_range = [doc["date"]]

    query = {
        "company_id": company_id,
        "status": {"$ne": "deleted"},
        "$or": strong_id_clauses + [{"date": {"$in": date_range}, "type": doc["type"]}],
    }
    if exclude_id:
        query["_id"] = {"$ne": exclude_id}

    candidates = list(db.transactions.find(query))
    desc_a = (doc.get("description") or "").strip().lower()

    matches = []
    for c in candidates:
        reasons = []
        if c["amount"] == doc["amount"] and c["date"] == doc["date"]:
            reasons.append("Same amount and date")
        if doc.get("reference_number") and c.get("reference_number") == doc["reference_number"]:
            reasons.append("Same reference number")
        if doc.get("invoice_number") and c.get("invoice_number") == doc["invoice_number"]:
            reasons.append("Same invoice number")
        if doc.get("document_id") and c.get("document_id") == doc["document_id"]:
            reasons.append("Linked to the same document")
        if doc.get("vendor_id") and c.get("vendor_id") == doc.get("vendor_id") and c["amount"] == doc["amount"]:
            reasons.append("Same vendor and amount")

        desc_b = (c.get("description") or "").strip().lower()
        if desc_a and desc_b:
            ratio = difflib.SequenceMatcher(None, desc_a, desc_b).ratio()
            if ratio >= DESCRIPTION_SIMILARITY_THRESHOLD:
                reasons.append("Very similar description")

        if reasons:
            matches.append({
                "id": str(c["_id"]), "reasons": reasons,
                "amount": c["amount"], "date": c["date"], "description": c.get("description", ""),
            })

    return matches


def create_transaction(user, data):
    db = mongo_connection.get_db()
    company, category_oid, vendor_oid, customer_oid, document_oid = _resolve_references(db, user, data)

    doc = {
        "date": data["date"].isoformat(),
        "amount": float(data["amount"]),
        "type": data["type"],
        "category_id": category_oid,
        "description": data.get("description", ""),
        "company_id": company["_id"],
        "vendor_id": vendor_oid,
        "customer_id": customer_oid,
        "payment_method": data.get("payment_method", "bank_transfer"),
        "reference_number": data.get("reference_number", ""),
        "invoice_number": data.get("invoice_number", ""),
        "gst_amount": float(data.get("gst_amount") or 0),
        "tds_amount": float(data.get("tds_amount") or 0),
        "document_id": document_oid,
        "status": data.get("status", "completed"),
        "tags": data.get("tags") or [],
        "notes": data.get("notes", ""),
        "created_by": str(user.id),
        "created_by_name": getattr(user, "full_name", ""),
        "created_at": now(),
        "updated_at": now(),
    }

    duplicates = find_possible_duplicates(db, company["_id"], doc)

    inserted = db.transactions.insert_one(doc)
    doc["_id"] = inserted.inserted_id
    _log_transaction_history(user, inserted.inserted_id, "created", {"amount": doc["amount"]})

    result = serialize(doc)
    # Kept for backward compatibility with existing frontend code.
    result["possible_duplicate_id"] = duplicates[0]["id"] if duplicates else None
    result["possible_duplicates"] = duplicates
    return result


def update_transaction(user, transaction_id, data):
    db = mongo_connection.get_db()
    transaction, company, is_owner = get_accessible_transaction(db, user, transaction_id)
    if not is_owner:
        raise PermissionDenied("Only the company owner can edit this transaction.")

    update_fields = {}
    for key, value in data.items():
        if key == "date":
            update_fields["date"] = value.isoformat()
        elif key in ("amount", "gst_amount", "tds_amount"):
            update_fields[key] = float(value)
        elif key == "category_id":
            update_fields["category_id"] = _validate_category(db, user, value)
        elif key == "vendor_id":
            update_fields["vendor_id"] = _validate_vendor(db, user, value) if value else None
        elif key == "customer_id":
            update_fields["customer_id"] = _validate_customer(db, user, value) if value else None
        elif key == "document_id":
            update_fields["document_id"] = _validate_document(db, company, value) if value else None
        elif key == "company_id":
            new_company, _ = get_accessible_company(db, user, value)
            update_fields["company_id"] = new_company["_id"]
        else:
            update_fields[key] = value

    update_fields["updated_at"] = now()
    db.transactions.update_one({"_id": transaction["_id"]}, {"$set": update_fields})
    _log_transaction_history(user, transaction["_id"], "updated", {"fields": list(data.keys())})
    return serialize(db.transactions.find_one({"_id": transaction["_id"]}))


def delete_transaction(user, transaction_id):
    db = mongo_connection.get_db()
    transaction, company, is_owner = get_accessible_transaction(db, user, transaction_id)
    if not is_owner:
        raise PermissionDenied("Only the company owner can delete this transaction.")
    db.transactions.update_one({"_id": transaction["_id"]}, {"$set": {"status": "deleted", "updated_at": now()}})
    _log_transaction_history(user, transaction["_id"], "deleted")


def bulk_delete_transactions(user, ids):
    results = []
    for transaction_id in ids:
        try:
            delete_transaction(user, transaction_id)
            results.append({"id": transaction_id, "success": True})
        except (NotFound, PermissionDenied) as exc:
            results.append({"id": transaction_id, "success": False, "error": str(exc)})
    return results


def bulk_update_transactions(user, ids, fields):
    db = mongo_connection.get_db()
    update_fields = {"updated_at": now()}
    if "status" in fields:
        update_fields["status"] = fields["status"]
    if "category_id" in fields:
        update_fields["category_id"] = _validate_category(db, user, fields["category_id"])
    if "tags" in fields:
        update_fields["tags"] = fields["tags"]

    results = []
    for transaction_id in ids:
        try:
            transaction, company, is_owner = get_accessible_transaction(db, user, transaction_id)
            if not is_owner:
                raise PermissionDenied("Only the company owner can edit this transaction.")
            per_transaction_fields = dict(update_fields)
            if "add_tags" in fields:
                existing_tags = transaction.get("tags") or []
                per_transaction_fields["tags"] = sorted(set(existing_tags) | set(fields["add_tags"]))
            db.transactions.update_one({"_id": transaction["_id"]}, {"$set": per_transaction_fields})
            _log_transaction_history(user, transaction["_id"], "bulk_updated", fields)
            results.append({"id": transaction_id, "success": True})
        except (NotFound, PermissionDenied) as exc:
            results.append({"id": transaction_id, "success": False, "error": str(exc)})
    return results


def list_transaction_tags(user):
    """Canonical tag vocabulary from the brief, plus any custom tags already
    in use on this user's transactions, so the UI can offer both.
    """
    from .serializers import TRANSACTION_TAGS

    db = mongo_connection.get_db()
    company_ids = get_user_company_ids(user)
    in_use = db.transactions.distinct("tags", {"company_id": {"$in": company_ids}})
    custom = sorted(t for t in in_use if t and t.lower() not in TRANSACTION_TAGS)
    return {"canonical": TRANSACTION_TAGS, "custom": custom}


def _log_transaction_history(user, transaction_id, action, metadata=None):
    db = mongo_connection.get_db()
    db.transaction_history.insert_one({
        "transaction_id": transaction_id,
        "user_id": str(user.id),
        "user_name": getattr(user, "full_name", ""),
        "action": action,
        "metadata": metadata or {},
        "created_at": now(),
    })
    log_activity(user, f"transaction_{action}", "transaction", transaction_id, metadata)


# ---------------------------------------------------------------------------
# Transactions - list / detail
# ---------------------------------------------------------------------------

CATEGORY_LOOKUP_CACHE_FIELDS = {"name": 1, "color": 1, "icon": 1, "type": 1}


def _enrich(db, items, company_ids, user=None):
    company_lookup = {str(c["_id"]): c["name"] for c in db.companies.find({"_id": {"$in": company_ids}}, {"name": 1})}
    category_ids = [c["category_id"] for c in items if c.get("category_id")]
    category_query = {"_id": {"$in": category_ids}}
    if user is not None:
        # Defense in depth: even though transactions should only ever
        # reference default/own categories (enforced at write time), never
        # let a lookup surface another user's private category here either.
        category_query = {"$and": [category_query, {"$or": [{"is_default": True}, {"owner_id": str(user.id)}]}]}
    category_lookup = {str(c["_id"]): c for c in db.transaction_categories.find(category_query)}
    vendor_ids = [c["vendor_id"] for c in items if c.get("vendor_id")]
    vendor_lookup = {str(v["_id"]): v["name"] for v in db.vendors.find({"_id": {"$in": vendor_ids}}, {"name": 1})}
    customer_ids = [c["customer_id"] for c in items if c.get("customer_id")]
    customer_lookup = {str(v["_id"]): v["name"] for v in db.customers.find({"_id": {"$in": customer_ids}}, {"name": 1})}

    for item in items:
        item["company_name"] = company_lookup.get(item["company_id"], "Unknown company")
        category = category_lookup.get(item.get("category_id"))
        item["category_name"] = category["name"] if category else None
        item["category_color"] = category["color"] if category else None
        item["vendor_name"] = vendor_lookup.get(item.get("vendor_id"))
        item["customer_name"] = customer_lookup.get(item.get("customer_id"))
    return items


def list_transactions(user, params):
    db = mongo_connection.get_db()
    company_ids = get_user_company_ids(user)

    company_filter = params.get("company_id")
    if company_filter:
        company, _ = get_accessible_company(db, user, company_filter)
        query = {"company_id": company["_id"]}
    else:
        query = {"company_id": {"$in": company_ids}}

    status_filter = params.get("status") or "all"
    if status_filter != "all":
        query["status"] = status_filter
    else:
        query["status"] = {"$ne": "deleted"}

    if params.get("type"):
        query["type"] = params["type"]
    if params.get("category_id"):
        query["category_id"] = to_object_id(params["category_id"], field="category_id")
    if params.get("vendor_id"):
        query["vendor_id"] = to_object_id(params["vendor_id"], field="vendor_id")
    if params.get("customer_id"):
        query["customer_id"] = to_object_id(params["customer_id"], field="customer_id")
    if params.get("ids"):
        raw_ids = params["ids"] if isinstance(params["ids"], list) else str(params["ids"]).split(",")
        query["_id"] = {"$in": [to_object_id(i.strip(), field="ids") for i in raw_ids if i.strip()]}
    if params.get("tag"):
        query["tags"] = params["tag"]

    date_from, date_to = params.get("date_from"), params.get("date_to")
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to
        query["date"] = date_query

    amount_min, amount_max = params.get("amount_min"), params.get("amount_max")
    if amount_min or amount_max:
        amount_query = {}
        if amount_min:
            amount_query["$gte"] = float(amount_min)
        if amount_max:
            amount_query["$lte"] = float(amount_max)
        query["amount"] = amount_query

    search = params.get("search")
    if search:
        query["$or"] = [
            {"description": {"$regex": search, "$options": "i"}},
            {"reference_number": {"$regex": search, "$options": "i"}},
            {"invoice_number": {"$regex": search, "$options": "i"}},
            {"notes": {"$regex": search, "$options": "i"}},
        ]

    sort_field = params.get("sort_by", "date")
    sort_dir = -1 if params.get("sort_dir", "desc") == "desc" else 1
    allowed_sort_fields = {"date", "amount", "created_at", "type", "status"}
    if sort_field not in allowed_sort_fields:
        sort_field = "date"

    result = paginate(
        db.transactions, query,
        page=params.get("page", 1), page_size=params.get("page_size", 10),
        sort_field=sort_field, sort_dir=sort_dir,
    )
    result["items"] = _enrich(db, result["items"], company_ids, user)
    return result


def get_transaction_detail(user, transaction_id):
    db = mongo_connection.get_db()
    transaction, company, is_owner = get_accessible_transaction(db, user, transaction_id)
    history = list(db.transaction_history.find({"transaction_id": transaction["_id"]}).sort("created_at", -1).limit(20))

    result = serialize(transaction)
    result = _enrich(db, [result], [transaction["company_id"]], user)[0]
    result["is_owner"] = is_owner
    result["history"] = serialize(history)
    return result


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

def get_dashboard_summary(user, company_id=None):
    db = mongo_connection.get_db()
    company_ids = get_user_company_ids(user)
    if company_id:
        company, _ = get_accessible_company(db, user, company_id)
        company_ids = [company["_id"]]

    base_query = {"company_id": {"$in": company_ids}, "status": {"$ne": "deleted"}}

    def sum_for(extra_query):
        agg = list(db.transactions.aggregate([
            {"$match": {**base_query, **extra_query}},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}}},
        ]))
        return agg[0]["total"] if agg else 0

    total_income = sum_for({"type": "income"})
    total_expenses = sum_for({"type": "expense"})

    from dateutil.relativedelta import relativedelta
    twelve_months_ago = (now() - relativedelta(months=11)).strftime("%Y-%m-01")

    monthly = list(db.transactions.aggregate([
        {"$match": {**base_query, "type": {"$in": ["income", "expense"]}, "date": {"$gte": twelve_months_ago}}},
        {"$group": {
            "_id": {"month": {"$substrCP": ["$date", 0, 7]}, "type": "$type"},
            "total": {"$sum": "$amount"},
        }},
        {"$sort": {"_id.month": 1}},
    ]))
    
    monthly_series: dict = {}
    monthly_income = 0.0
    monthly_expenses = 0.0
    this_month_prefix = now().strftime("%Y-%m")
    
    for row in monthly:
        month = row["_id"]["month"]
        txn_type = row["_id"]["type"]
        total = row["total"]
        
        monthly_series.setdefault(month, {"month": month, "income": 0, "expense": 0})
        monthly_series[month][txn_type] = total
        
        if month == this_month_prefix:
            if txn_type == "income":
                monthly_income = total
            elif txn_type == "expense":
                monthly_expenses = total
                
    monthly_trend = list(monthly_series.values())[-12:]

    category_breakdown = list(db.transactions.aggregate([
        {"$match": {**base_query, "type": "expense"}},
        {"$group": {"_id": "$category_id", "total": {"$sum": "$amount"}}},
        {"$sort": {"total": -1}},
        {"$limit": 8},
    ]))
    category_lookup = {
        c["_id"]: c for c in db.transaction_categories.find({"$or": [{"is_default": True}, {"owner_id": str(user.id)}]})
    }
    expense_breakdown = [
        {
            "category_id": str(row["_id"]) if row["_id"] else None,
            "category_name": category_lookup.get(row["_id"], {}).get("name", "Uncategorized"),
            "color": category_lookup.get(row["_id"], {}).get("color", "#94A3B8"),
            "total": row["total"],
        }
        for row in category_breakdown
    ]

    recent = list(db.transactions.find(base_query).sort("created_at", -1).limit(8))
    recent = _enrich(db, [serialize(r) for r in recent], company_ids, user)

    today_str = now().strftime("%Y-%m-%d")
    today_income = sum_for({"type": "income", "date": today_str})
    today_expenses = sum_for({"type": "expense", "date": today_str})
    pending_transactions_count = db.transactions.count_documents({**base_query, "status": "pending"})

    # --- Financial Intelligence -------------------------------------------

    def _largest(txn_type):
        doc = db.transactions.find_one({**base_query, "type": txn_type}, sort=[("amount", -1)])
        return _enrich(db, [serialize(doc)], company_ids, user)[0] if doc else None

    largest_expense = _largest("expense")
    largest_income = _largest("income")

    def _top_party(field, collection):
        rows = list(db.transactions.aggregate([
            {"$match": {**base_query, field: {"$ne": None}}},
            {"$group": {"_id": f"${field}", "total": {"$sum": "$amount"}, "count": {"$sum": 1}}},
            {"$sort": {"total": -1}},
            {"$limit": 5},
        ]))
        names = {p["_id"]: p["name"] for p in collection.find({"_id": {"$in": [r["_id"] for r in rows]}}, {"name": 1})}
        return [{"id": str(r["_id"]), "name": names.get(r["_id"], "Unknown"), "total": r["total"], "count": r["count"]} for r in rows]

    top_vendors = _top_party("vendor_id", db.vendors)
    top_customers = _top_party("customer_id", db.customers)

    def _growth(key):
        if len(monthly_trend) < 2:
            return 0
        previous, current = monthly_trend[-2][key], monthly_trend[-1][key]
        if previous == 0:
            return 100.0 if current > 0 else 0.0
        return round((current - previous) / previous * 100, 2)

    income_growth = _growth("income")
    expense_growth = _growth("expense")

    savings_trend = [{"month": m["month"], "savings": round(m["income"] - m["expense"], 2)} for m in monthly_trend]
    running_cash_flow = 0
    cash_flow_trend = []
    for m in monthly_trend:
        running_cash_flow += m["income"] - m["expense"]
        cash_flow_trend.append({"month": m["month"], "cash_flow": round(running_cash_flow, 2)})

    total_transactions_count = db.transactions.count_documents(base_query)
    active_vendor_ids = [v for v in db.transactions.distinct("vendor_id", base_query) if v]
    active_customer_ids = [c for c in db.transactions.distinct("customer_id", base_query) if c]
    financial_kpis = {
        "total_transactions": total_transactions_count,
        "avg_transaction_value": round((total_income + total_expenses) / total_transactions_count, 2) if total_transactions_count else 0,
        "active_vendors": len(active_vendor_ids),
        "active_customers": len(active_customer_ids),
    }

    def _recent_party_names(field, collection, limit=5):
        recent_ids = []
        for doc in db.transactions.find(base_query, {field: 1}).sort("created_at", -1).limit(200):
            val = doc.get(field)
            if val is not None and val not in recent_ids:
                recent_ids.append(val)
            if len(recent_ids) >= limit:
                break
        names = {p["_id"]: p["name"] for p in collection.find({"_id": {"$in": recent_ids}}, {"name": 1})}
        return [{"id": str(pid), "name": names.get(pid, "Unknown")} for pid in recent_ids]

    recent_vendors = _recent_party_names("vendor_id", db.vendors)
    recent_customers = _recent_party_names("customer_id", db.customers)

    return {
        "total_income": total_income,
        "total_expenses": total_expenses,
        "net_profit": total_income - total_expenses,
        "monthly_income": monthly_income,
        "monthly_expenses": monthly_expenses,
        "monthly_trend": monthly_trend,
        "expense_breakdown": expense_breakdown,
        "recent_transactions": recent,
        "today_income": today_income,
        "today_expenses": today_expenses,
        "pending_transactions_count": pending_transactions_count,
        "recent_vendors": recent_vendors,
        "recent_customers": recent_customers,
        "financial_intelligence": {
            "largest_expense": largest_expense,
            "largest_income": largest_income,
            "top_vendors": top_vendors,
            "top_customers": top_customers,
            "income_growth": income_growth,
            "expense_growth": expense_growth,
            "savings_trend": savings_trend,
            "cash_flow_trend": cash_flow_trend,
            "kpis": financial_kpis,
        },
    }


# ---------------------------------------------------------------------------
# Payment method analytics
# ---------------------------------------------------------------------------

def get_payment_analytics(user, company_id=None, date_from=None, date_to=None):
    db = mongo_connection.get_db()
    company_ids = get_user_company_ids(user)
    if company_id:
        company, _ = get_accessible_company(db, user, company_id)
        company_ids = [company["_id"]]

    base_query = {"company_id": {"$in": company_ids}, "status": {"$ne": "deleted"}}
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to
        base_query["date"] = date_query

    by_method = list(db.transactions.aggregate([
        {"$match": base_query},
        {"$group": {"_id": "$payment_method", "total_amount": {"$sum": "$amount"}, "count": {"$sum": 1}}},
        {"$sort": {"total_amount": -1}},
    ]))
    total_amount_all = sum(row["total_amount"] for row in by_method) or 1
    total_count_all = sum(row["count"] for row in by_method)

    methods = [
        {
            "payment_method": row["_id"] or "other",
            "total_amount": row["total_amount"],
            "count": row["count"],
            "usage_percentage": round(row["total_amount"] / total_amount_all * 100, 2),
        }
        for row in by_method
    ]
    preferred_payment_method = methods[0]["payment_method"] if methods else None

    monthly_raw = list(db.transactions.aggregate([
        {"$match": base_query},
        {"$group": {
            "_id": {"month": {"$substrCP": ["$date", 0, 7]}, "method": "$payment_method"},
            "total": {"$sum": "$amount"},
        }},
        {"$sort": {"_id.month": 1}},
    ]))
    monthly_series: dict = {}
    all_methods = sorted({row["_id"]["method"] or "other" for row in monthly_raw})
    for row in monthly_raw:
        month = row["_id"]["month"]
        entry = monthly_series.setdefault(month, {"month": month, **{m: 0 for m in all_methods}})
        entry[row["_id"]["method"] or "other"] = row["total"]
    monthly_trend = list(monthly_series.values())[-12:]

    return {
        "total_transactions": total_count_all,
        "total_amount": total_amount_all if by_method else 0,
        "preferred_payment_method": preferred_payment_method,
        "methods": methods,
        "monthly_trend": monthly_trend,
    }


# ---------------------------------------------------------------------------
# Recurring transaction detection
#
# Read-only intelligence: scans existing transactions and writes findings to
# a *separate* `recurring_patterns` collection. Never touches a transaction
# document itself, per the brief.
# ---------------------------------------------------------------------------

RECURRING_MIN_OCCURRENCES = 3
RECURRING_AMOUNT_TOLERANCE = 0.15  # 15% variance still counts as "the same" payment
RECURRING_MIN_GAP_DAYS = 20
RECURRING_MAX_GAP_DAYS = 40

PATTERN_KEYWORDS = [
    ("salary", ["salary", "payroll", "stipend"]),
    ("rent", ["rent"]),
    ("home_loan_emi", ["home loan", "housing loan", "home emi"]),
    ("education_loan_emi", ["education loan", "student loan"]),
    ("vehicle_loan_emi", ["vehicle loan", "car loan", "auto loan"]),
    ("personal_loan_emi", ["personal loan"]),
    ("insurance_premium", ["insurance", "premium", "lic"]),
    ("sip", ["sip", "mutual fund", "systematic investment"]),
    ("investment", ["investment", "invest", "stocks", "shares"]),
    ("utility_bill", ["electricity", "water bill", "utility", "gas bill", "broadband"]),
    ("subscription", ["subscription", "netflix", "prime", "spotify", "saas"]),
]


def _normalize_description(text):
    text = (text or "").lower().strip()
    text = re.sub(r"\d+", "", text)          # drop invoice/ref numbers embedded in text
    text = re.sub(r"[^a-z\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _classify_recurring_pattern(description, category_name, tags):
    text = f"{description or ''} {category_name or ''}".lower()
    if "recurring" in [t.lower() for t in (tags or [])] and not text.strip():
        return "other_recurring"
    for pattern_type, keywords in PATTERN_KEYWORDS:
        if any(kw in text for kw in keywords):
            return pattern_type
    return "other_recurring"


def detect_recurring_transactions(user, company_id=None):
    """Groups transactions by (company, type, vendor-or-description) and
    flags groups that recur roughly monthly with a stable amount. Persists
    results to `recurring_patterns`; does not modify `transactions`.
    """
    db = mongo_connection.get_db()
    if company_id:
        company, _ = get_accessible_company(db, user, company_id)
        company_ids = [company["_id"]]
    else:
        company_ids = get_user_company_ids(user)

    transactions = list(db.transactions.find({"company_id": {"$in": company_ids}, "status": {"$ne": "deleted"}}).sort("date", -1).limit(2000))
    categories = {c["_id"]: c for c in db.transaction_categories.find()}

    groups = defaultdict(list)
    for t in transactions:
        group_key = (
            str(t["company_id"]), t["type"],
            str(t["vendor_id"]) if t.get("vendor_id") else _normalize_description(t.get("description", "")),
        )
        groups[group_key].append(t)

    patterns = []
    for txs in groups.values():
        if len(txs) < RECURRING_MIN_OCCURRENCES:
            continue

        txs = sorted(txs, key=lambda t: t["date"])
        amounts = [t["amount"] for t in txs]
        avg_amount = sum(amounts) / len(amounts)
        if avg_amount <= 0:
            continue
        if any(abs(a - avg_amount) / avg_amount > RECURRING_AMOUNT_TOLERANCE for a in amounts):
            continue

        try:
            dates = [date_cls.fromisoformat(t["date"]) for t in txs]
        except ValueError:
            continue
        gaps = [(dates[i + 1] - dates[i]).days for i in range(len(dates) - 1)]
        recent_gaps = gaps[-3:] if len(gaps) >= 3 else gaps
        if not recent_gaps or not all(RECURRING_MIN_GAP_DAYS <= g <= RECURRING_MAX_GAP_DAYS for g in recent_gaps):
            continue

        last = txs[-1]
        category = categories.get(last.get("category_id"), {})
        pattern_type = _classify_recurring_pattern(last.get("description"), category.get("name"), last.get("tags"))
        amount_spread = max(abs(a - avg_amount) for a in amounts) / avg_amount if avg_amount else 1
        confidence = round(max(0.0, min(1.0, (len(txs) / 6) * (1 - amount_spread))), 2)

        patterns.append({
            "company_id": last["company_id"],
            "owner_id": str(user.id),
            "pattern_type": pattern_type,
            "vendor_id": last.get("vendor_id"),
            "category_id": last.get("category_id"),
            "type": last["type"],
            "average_amount": round(avg_amount, 2),
            "occurrences": len(txs),
            "last_date": last["date"],
            "next_expected_date": (dates[-1] + timedelta(days=30)).isoformat(),
            "confidence": confidence,
            "transaction_ids": [t["_id"] for t in txs],
            "detected_at": now(),
        })

    db.recurring_patterns.delete_many({"company_id": {"$in": company_ids}, "owner_id": str(user.id)})
    if patterns:
        db.recurring_patterns.insert_many(patterns)
    return _enrich_recurring_patterns(db, serialize(patterns))


def list_recurring_patterns(user, company_id=None):
    db = mongo_connection.get_db()
    if company_id:
        company, _ = get_accessible_company(db, user, company_id)
        company_ids = [company["_id"]]
    else:
        company_ids = get_user_company_ids(user)

    docs = list(db.recurring_patterns.find({
        "company_id": {"$in": company_ids}, "owner_id": str(user.id),
    }).sort("confidence", -1))
    return _enrich_recurring_patterns(db, serialize(docs))


def _enrich_recurring_patterns(db, patterns):
    vendor_ids = [to_object_id(p["vendor_id"], field="vendor_id") for p in patterns if p.get("vendor_id")]
    category_ids = [to_object_id(p["category_id"], field="category_id") for p in patterns if p.get("category_id")]
    vendor_lookup = {str(v["_id"]): v["name"] for v in db.vendors.find({"_id": {"$in": vendor_ids}}, {"name": 1})}
    category_lookup = {str(c["_id"]): c["name"] for c in db.transaction_categories.find({"_id": {"$in": category_ids}}, {"name": 1})}
    for p in patterns:
        p["vendor_name"] = vendor_lookup.get(p.get("vendor_id"))
        p["category_name"] = category_lookup.get(p.get("category_id"))
    return patterns


# ---------------------------------------------------------------------------
# CSV import / export
# ---------------------------------------------------------------------------

CSV_COLUMNS = [
    "date", "amount", "type", "category", "description", "vendor", "customer",
    "payment_method", "reference_number", "invoice_number", "gst_amount", "tds_amount", "status", "notes",
]


def export_transactions_csv(user, params):
    result = list_transactions(user, {**params, "page": 1, "page_size": 10000})
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(CSV_COLUMNS)
    for row in result["items"]:
        writer.writerow([
            row["date"], row["amount"], row["type"], row.get("category_name") or "",
            row.get("description", ""), row.get("vendor_name") or "", row.get("customer_name") or "",
            row.get("payment_method", ""), row.get("reference_number", ""), row.get("invoice_number", ""),
            row.get("gst_amount", 0), row.get("tds_amount", 0), row.get("status", ""), row.get("notes", ""),
        ])
    return buffer.getvalue()


def _rows_for_export(user, params):
    """Shared row-extraction used by both the CSV and Excel exporters."""
    result = list_transactions(user, {**params, "page": 1, "page_size": 10000})
    for row in result["items"]:
        yield [
            row["date"], row["amount"], row["type"], row.get("category_name") or "",
            row.get("description", ""), row.get("vendor_name") or "", row.get("customer_name") or "",
            row.get("payment_method", ""), row.get("reference_number", ""), row.get("invoice_number", ""),
            row.get("gst_amount", 0), row.get("tds_amount", 0), row.get("status", ""), row.get("notes", ""),
        ]


def export_transactions_xlsx(user, params):
    """Excel counterpart of `export_transactions_csv`. Returns raw xlsx bytes."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transactions"

    sheet.append(CSV_COLUMNS)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for row in _rows_for_export(user, params):
        sheet.append(row)

    for column_cells in sheet.columns:
        length = max((len(str(c.value)) for c in column_cells if c.value is not None), default=10)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 40)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_import_error_report_xlsx(results):
    """Builds a downloadable Excel error report from an import's per-row
    results (the same `results` list already returned in the JSON response).
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Import Errors"
    sheet.append(["Row", "Status", "Error"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for r in results:
        if not r["success"]:
            sheet.append([r["row"], "Failed", r.get("error", "")])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _lookup_maps(db, user):
    """Shared name->id lookups used by every import format."""
    return (
        {c["name"].lower(): c["_id"] for c in db.transaction_categories.find({"$or": [{"is_default": True}, {"owner_id": str(user.id)}]})},
        {v["name"].lower(): v["_id"] for v in db.vendors.find({"owner_id": str(user.id)})},
        {c["name"].lower(): c["_id"] for c in db.customers.find({"owner_id": str(user.id)})},
    )


def _row_to_transaction_doc(row, user, company_id, categories_by_name, vendors_by_name, customers_by_name):
    """Build a transaction document from a single imported row (CSV or Excel).

    `row` is a plain dict keyed by the CSV_COLUMNS header names, with all
    values already coerced to strings/numbers as read from the source file.
    Raises ValidationError/KeyError/ValueError on bad data — callers turn
    that into a per-row error-report entry instead of failing the batch.
    """
    category_id = categories_by_name.get((row.get("category") or "").strip().lower())
    if not category_id:
        raise ValidationError(f"Unknown category '{row.get('category')}'.")

    return {
        "date": str(row["date"]).strip(),
        "amount": float(row["amount"]),
        "type": str(row["type"]).strip().lower(),
        "category_id": category_id,
        "description": row.get("description", "") or "",
        "company_id": company_id,
        "vendor_id": vendors_by_name.get((row.get("vendor") or "").strip().lower()),
        "customer_id": customers_by_name.get((row.get("customer") or "").strip().lower()),
        "payment_method": (row.get("payment_method") or "bank_transfer").strip().lower(),
        "reference_number": row.get("reference_number", "") or "",
        "invoice_number": row.get("invoice_number", "") or "",
        "gst_amount": float(row.get("gst_amount") or 0),
        "tds_amount": float(row.get("tds_amount") or 0),
        "document_id": None,
        "status": (row.get("status") or "completed").strip().lower(),
        "tags": [],
        "notes": row.get("notes", "") or "",
        "created_by": str(user.id),
        "created_by_name": getattr(user, "full_name", ""),
        "created_at": now(),
        "updated_at": now(),
    }


def _run_import(user, company_id, rows_with_numbers):
    """Insert transactions for an iterable of (row_number, row_dict), building
    a per-row error report. Shared by the CSV and Excel importers.
    """
    db = mongo_connection.get_db()
    company, _ = get_accessible_company(db, user, company_id)
    categories_by_name, vendors_by_name, customers_by_name = _lookup_maps(db, user)

    results = []
    for row_number, row in rows_with_numbers:
        try:
            doc = _row_to_transaction_doc(row, user, company["_id"], categories_by_name, vendors_by_name, customers_by_name)
            inserted = db.transactions.insert_one(doc)
            _log_transaction_history(user, inserted.inserted_id, "imported")
            results.append({"row": row_number, "success": True})
        except (KeyError, ValueError, ValidationError) as exc:
            results.append({"row": row_number, "success": False, "error": str(exc)})
    return results


def import_transactions_csv(user, company_id, file_obj):
    text = file_obj.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows_with_numbers = ((row_number, row) for row_number, row in enumerate(reader, start=2))
    return _run_import(user, company_id, rows_with_numbers)


def import_transactions_xlsx(user, company_id, file_obj):
    """Excel counterpart of `import_transactions_csv`. Column order in the
    sheet doesn't matter — the header row is matched by name against
    CSV_COLUMNS (a simple, order-independent form of "column mapping"),
    same as the CSV DictReader does for the CSV path.
    """
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    sheet = workbook.active

    rows = sheet.iter_rows(values_only=True)
    try:
        header = [str(h).strip().lower() if h is not None else "" for h in next(rows)]
    except StopIteration:
        raise ValidationError("The uploaded Excel file has no header row.")

    def rows_with_numbers():
        for offset, values in enumerate(rows, start=2):
            if values is None or all(v is None for v in values):
                continue  # skip blank rows
            row = dict(zip(header, values))
            # Normalise everything to strings the same way CSV rows arrive as,
            # except leave amount/gst_amount/tds_amount numeric-friendly.
            row = {k: ("" if v is None else v) for k, v in row.items()}
            yield offset, row

    return _run_import(user, company_id, rows_with_numbers())


def import_transactions_file(user, company_id, file_obj, filename):
    """Dispatches to the CSV or Excel importer based on file extension."""
    lower_name = (filename or "").lower()
    if lower_name.endswith((".xlsx", ".xls")):
        return import_transactions_xlsx(user, company_id, file_obj)
    return import_transactions_csv(user, company_id, file_obj)


# ---------------------------------------------------------------------------
# Budget management
#
# A budget is a target for a company + month (+ optional category). Actual
# spend is always computed live from `transactions` — never duplicated or
# cached — so it's automatically correct as transactions change.
# ---------------------------------------------------------------------------

MONTH_REGEX = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")


def _validate_month(month):
    if not month or not MONTH_REGEX.match(month):
        raise ValidationError({"month": "Provide a month as YYYY-MM."})
    return month


def _actual_spend(db, company_id, month, category_id=None):
    query = {
        "company_id": company_id, "type": "expense", "status": {"$ne": "deleted"},
        "date": {"$gte": f"{month}-01", "$lte": f"{month}-31"},
    }
    if category_id:
        query["category_id"] = category_id
    agg = list(db.transactions.aggregate([
        {"$match": query},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}},
    ]))
    return agg[0]["total"] if agg else 0


def set_budget(user, data):
    """Creates or updates (upserts) the budget for a company + month +
    optional category — one target per combination, so re-submitting the
    same form just adjusts the existing budget instead of duplicating it.
    """
    db = mongo_connection.get_db()
    company, _ = get_accessible_company(db, user, data["company_id"])
    month = _validate_month(data["month"])
    category_oid = _validate_category(db, user, data["category_id"]) if data.get("category_id") else None

    key = {"owner_id": str(user.id), "company_id": company["_id"], "month": month, "category_id": category_oid}
    db.budgets.update_one(
        key,
        {"$set": {**key, "amount": float(data["amount"]), "updated_at": now()}, "$setOnInsert": {"created_at": now()}},
        upsert=True,
    )
    doc = db.budgets.find_one(key)
    return _enrich_budget(db, serialize(doc), user)


def list_budgets(user, company_id, month):
    db = mongo_connection.get_db()
    company, _ = get_accessible_company(db, user, company_id)
    month = _validate_month(month)
    docs = list(db.budgets.find({"owner_id": str(user.id), "company_id": company["_id"], "month": month}))
    return [_enrich_budget(db, serialize(d), user) for d in docs]


def _enrich_budget(db, budget, user):
    category_oid = to_object_id(budget["category_id"], field="category_id") if budget.get("category_id") else None
    company_oid = to_object_id(budget["company_id"], field="company_id")

    category_name = "Overall (all categories)"
    if category_oid:
        category = db.transaction_categories.find_one({"_id": category_oid})
        category_name = category["name"] if category else "Unknown category"

    actual = _actual_spend(db, company_oid, budget["month"], category_oid)
    remaining = budget["amount"] - actual
    progress_pct = round(min(actual / budget["amount"] * 100, 999), 1) if budget["amount"] else 0

    budget["category_name"] = category_name
    budget["actual_spent"] = actual
    budget["remaining"] = remaining
    budget["progress_percentage"] = progress_pct
    budget["is_overspending"] = actual > budget["amount"]
    return budget


def delete_budget(user, budget_id):
    db = mongo_connection.get_db()
    oid = to_object_id(budget_id, field="budget_id")
    budget = get_or_404(db.budgets, {"_id": oid}, "Budget not found.")
    if budget.get("owner_id") != str(user.id):
        raise PermissionDenied("You can only delete your own budgets.")
    db.budgets.delete_one({"_id": oid})


def get_budget_summary(user, company_id, month):
    """Budget dashboard: totals, per-category breakdown, and overspending
    alerts for a company + month, built from `list_budgets`."""
    budgets = list_budgets(user, company_id, month)
    total_budgeted = sum(b["amount"] for b in budgets)
    total_actual = sum(b["actual_spent"] for b in budgets)
    overspending = [b for b in budgets if b["is_overspending"]]

    return {
        "month": month,
        "total_budgeted": total_budgeted,
        "total_actual": total_actual,
        "total_remaining": total_budgeted - total_actual,
        "overall_progress_percentage": round(min(total_actual / total_budgeted * 100, 999), 1) if total_budgeted else 0,
        "budgets": budgets,
        "overspending_alerts": [
            {"budget_id": b["id"], "category_name": b["category_name"], "amount": b["amount"], "actual_spent": b["actual_spent"]}
            for b in overspending
        ],
    }
